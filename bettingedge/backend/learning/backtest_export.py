"""
Export multi-feuilles Excel d'un BacktestResult.

Feuilles produites :
  - Summary       : KPIs globaux + paramètres + interprétation
  - Bets          : tous les paris simulés (1 ligne par pari)
  - By niche      : breakdown ROI/hit_rate par niche
  - By league     : breakdown par ligue
  - Niche × League: matrice croisée
  - EV bins       : performance par tranche d'EV (insights critiques)
  - Odds bins     : performance par tranche de cote
  - Cumulative    : courbe cumulée du profit (chronologique)

Utilisable via /api/backtest/export.xlsx (téléchargement direct) pour
analyse externe (autre IA, Excel, Power BI...).
"""
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


def _style_header(ws, row=1):
    bold = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="2C3E50")
    for cell in ws[row]:
        cell.font = bold
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")


def _autosize(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)


def _df_to_sheet(wb: Workbook, name: str, df: pd.DataFrame, freeze: bool = True):
    ws = wb.create_sheet(name)
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    _style_header(ws)
    if freeze:
        ws.freeze_panes = "A2"
    _autosize(ws)
    return ws


def export_to_xlsx(result, params: dict | None = None) -> bytes:
    """Construit un .xlsx à partir d'un BacktestResult. Renvoie les bytes."""
    if not result.bets:
        raise ValueError("Aucun pari dans le résultat — lance un backtest d'abord")

    df = pd.DataFrame([{
        "league":      b.league,
        "date":        b.date,
        "home":        b.home,
        "away":        b.away,
        "niche":       b.niche,
        "p_estimated": b.p_estimated,
        "odds_taken":  b.odds_taken,
        "odds_fair":   b.odds_fair,
        "ev_pct":      round(b.ev_expected * 100, 3),
        "stake":       b.stake,
        "won":         b.won,
        "profit":      b.profit,
    } for b in result.bets])
    # Date parseable pour tri chronologique
    df["date_dt"] = pd.to_datetime(df["date"], format="%d/%m/%Y", errors="coerce")

    wb = Workbook()
    # Vire la feuille par défaut
    ws_default = wb.active
    wb.remove(ws_default)

    # ── Summary ────────────────────────────────────────────────────────────
    s = result.summary()
    summary_rows = [
        ("Métrique", "Valeur"),
        ("Paris simulés",       s["n_bets"]),
        ("Paris gagnés",        s["n_won"]),
        ("Hit rate",            f"{s['hit_rate_pct']}%"),
        ("Mise totale",         f"{s['total_stake']} €"),
        ("Profit total",        f"{s['total_profit']} €"),
        ("ROI global",          f"{s['roi_pct']}%"),
        ("Verdict objectif (>5%)", "✓ atteint" if s["roi"] >= 0.05 else "✗ sous l'objectif"),
        ("", ""),
        ("Paramètres backtest", ""),
    ]
    if params:
        for k, v in params.items():
            summary_rows.append((k, str(v)))
    summary_rows.extend([
        ("", ""),
        ("Note importante", "In-sample : DC calibré sur ces matchs → ROI optimiste."),
        ("", "Walk-forward (validation OOS) reste à implémenter."),
    ])
    ws = wb.create_sheet("Summary")
    for r in summary_rows:
        ws.append(list(r))
    _style_header(ws)
    _autosize(ws)

    # ── By niche ──────────────────────────────────────────────────────────
    niche_df = pd.DataFrame([
        {"niche": k, **{kk: vv for kk, vv in v.items()}}
        for k, v in s["by_niche"].items()
    ]).sort_values("roi", ascending=False) if s["by_niche"] else pd.DataFrame()
    _df_to_sheet(wb, "By niche", niche_df)

    # ── By league ─────────────────────────────────────────────────────────
    league_df = pd.DataFrame([
        {"league": k, **{kk: vv for kk, vv in v.items()}}
        for k, v in s["by_league"].items()
    ]).sort_values("roi", ascending=False) if s["by_league"] else pd.DataFrame()
    _df_to_sheet(wb, "By league", league_df)

    # ── Niche × League ────────────────────────────────────────────────────
    cross = df.groupby(["niche", "league"], observed=True).agg(
        n=("won", "count"), won=("won", "sum"),
        stake=("stake", "sum"), profit=("profit", "sum"),
    ).reset_index()
    cross["hit_rate_pct"] = (cross["won"] / cross["n"] * 100).round(1)
    cross["roi_pct"] = (cross["profit"] / cross["stake"] * 100).round(2)
    cross = cross.sort_values("roi_pct", ascending=False)
    _df_to_sheet(wb, "Niche x League", cross)

    # ── EV bins ───────────────────────────────────────────────────────────
    df["ev_bin"] = pd.cut(df["ev_pct"], bins=[0, 2, 5, 10, 15, 25, 50, 100],
                          labels=["0-2%", "2-5%", "5-10%", "10-15%", "15-25%", "25-50%", "50%+"])
    ev_bin = df.groupby("ev_bin", observed=True).agg(
        n=("won", "count"), won=("won", "sum"),
        stake=("stake", "sum"), profit=("profit", "sum"),
    ).reset_index()
    ev_bin["hit_rate_pct"] = (ev_bin["won"] / ev_bin["n"] * 100).round(1)
    ev_bin["roi_pct"] = (ev_bin["profit"] / ev_bin["stake"] * 100).round(2)
    _df_to_sheet(wb, "EV bins", ev_bin)

    # ── Odds bins ─────────────────────────────────────────────────────────
    df["odds_bin"] = pd.cut(df["odds_taken"], bins=[1, 1.5, 2, 2.5, 3.5, 5, 10, 50],
                            labels=["1.0-1.5", "1.5-2.0", "2.0-2.5", "2.5-3.5", "3.5-5", "5-10", "10+"])
    odds_bin = df.groupby("odds_bin", observed=True).agg(
        n=("won", "count"), won=("won", "sum"),
        stake=("stake", "sum"), profit=("profit", "sum"),
    ).reset_index()
    odds_bin["hit_rate_pct"] = (odds_bin["won"] / odds_bin["n"] * 100).round(1)
    odds_bin["roi_pct"] = (odds_bin["profit"] / odds_bin["stake"] * 100).round(2)
    _df_to_sheet(wb, "Odds bins", odds_bin)

    # ── Cumulative (courbe profit chronologique) ─────────────────────────
    chrono = df.dropna(subset=["date_dt"]).sort_values("date_dt").reset_index(drop=True)
    chrono["cumulative_profit"] = chrono["profit"].cumsum().round(2)
    chrono["peak"] = chrono["cumulative_profit"].cummax()
    chrono["drawdown"] = (chrono["peak"] - chrono["cumulative_profit"]).round(2)
    cum_export = chrono[["date", "league", "niche", "home", "away", "odds_taken",
                         "ev_pct", "won", "profit", "cumulative_profit", "drawdown"]]
    _df_to_sheet(wb, "Cumulative", cum_export)

    # ── Bets (toutes les transactions, en dernier pour ne pas être feuille active) ──
    df_export = df.drop(columns=["date_dt", "ev_bin", "odds_bin"], errors="ignore")
    _df_to_sheet(wb, "Bets", df_export)

    # Active la feuille Summary à l'ouverture
    wb.active = wb.sheetnames.index("Summary")

    # Serialise en bytes
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
